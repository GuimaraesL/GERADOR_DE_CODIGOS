import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import json

def carregar_codigos_existentes(filepath: str, aba='aba1') -> pd.DataFrame:
    wb = load_workbook(filepath)
    ws = wb[aba]
    dados_a, dados_b = [], []

    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        dados_a.append(row[0] if row[0] else None)
        dados_b.append(row[1] if len(row) > 1 and row[1] else None)

    df = pd.DataFrame({'Codigo': dados_a + dados_b}).dropna()
    with open("codigos.json", "w") as f:
        json.dump(df.to_dict(orient="records"), f)
    return df

def extrair_siglas(filepath: str, aba: str, coluna: int) -> pd.DataFrame:
    wb = load_workbook(filepath)
    ws = wb[aba]
    siglas = [row[coluna - 1].value for row in ws.iter_rows(min_row=1, min_col=coluna, max_col=coluna)]
    return pd.DataFrame(siglas, columns=["Sigla"])

def salvar_resultado(filepath: str, aba: str, df_resultado: pd.DataFrame):
    wb = load_workbook(filepath)
    original_aba = aba
    counter = 1
    while aba in wb.sheetnames:
        aba = f"{original_aba}{counter}"
        counter += 1
    ws = wb.create_sheet(aba)

    for r in dataframe_to_rows(df_resultado, index=False, header=True):
        ws.append(r)

    wb.save(filepath)